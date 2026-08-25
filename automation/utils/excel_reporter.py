import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from automation.config.config import Config

class ExcelReporter:
    """Generates enterprise Excel reports for Selenium test execution results."""

    def __init__(self, results):
        self.results = results
        Config.initialize_directories()

    def generate_all_excel_reports(self):
        self._generate_main_automation_report()
        self._generate_passed_tests_report()
        self._generate_failed_tests_report()
        self._generate_summary_report()

    def _get_styles(self):
        font_title = Font(name="Calibri", size=16, bold=True, color="1F4E78")
        font_section = Font(name="Calibri", size=12, bold=True, color="1F4E78")
        font_header = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        font_bold = Font(name="Calibri", size=11, bold=True)
        font_regular = Font(name="Calibri", size=10)
        
        fill_navy = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        fill_pass = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        fill_fail = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        fill_skip = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        
        border_thin = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        return {
            "title": font_title, "section": font_section, "header": font_header,
            "bold": font_bold, "regular": font_regular,
            "fill_navy": fill_navy, "fill_pass": fill_pass, "fill_fail": fill_fail, "fill_skip": fill_skip,
            "border": border_thin
        }

    def _generate_main_automation_report(self):
        wb = openpyxl.Workbook()
        s = self._get_styles()

        # Sheet 1: Executed Test Cases
        ws1 = wb.active
        ws1.title = "Executed Test Cases"
        ws1.views.sheetView[0].showGridLines = True

        headers = ["Test ID", "Module", "Test Name", "Status", "Execution Time (ms)", "Priority", "Expected Result", "Actual Result"]
        ws1.append(headers)

        for col_num, h_text in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col_num)
            cell.font = s["header"]
            cell.fill = s["fill_navy"]
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for r_idx, res in enumerate(self.results, 2):
            ws1.append([
                res["test_id"], res["module"], res["name"], res["status"],
                res["exec_time_ms"], res["priority"], res["expected"], res["actual"]
            ])

            status_cell = ws1.cell(row=r_idx, column=4)
            if res["status"] == "PASS":
                status_cell.fill = s["fill_pass"]
                status_cell.font = Font(name="Calibri", size=10, bold=True, color="276A3C")
            elif res["status"] == "FAIL":
                status_cell.fill = s["fill_fail"]
                status_cell.font = Font(name="Calibri", size=10, bold=True, color="9C0006")
            else:
                status_cell.fill = s["fill_skip"]
                status_cell.font = Font(name="Calibri", size=10, bold=True, color="B7791F")

            for c_idx in range(1, 9):
                ws1.cell(row=r_idx, column=c_idx).border = s["border"]

        # Sheet 2: Passed Tests
        ws2 = wb.create_sheet(title="Passed Tests")
        ws2.views.sheetView[0].showGridLines = True
        ws2.append(headers)
        for col_num, h_text in enumerate(headers, 1):
            ws2.cell(row=1, column=col_num).font = s["header"]
            ws2.cell(row=1, column=col_num).fill = s["fill_navy"]
        for res in self.results:
            if res["status"] == "PASS":
                ws2.append([res["test_id"], res["module"], res["name"], res["status"], res["exec_time_ms"], res["priority"], res["expected"], res["actual"]])

        # Sheet 3: Failed Tests
        ws3 = wb.create_sheet(title="Failed Tests")
        ws3.views.sheetView[0].showGridLines = True
        ws3.append(headers + ["Failure Reason", "Screenshot Path"])
        for col_num, h_text in enumerate(headers + ["Failure Reason", "Screenshot Path"], 1):
            ws3.cell(row=1, column=col_num).font = s["header"]
            ws3.cell(row=1, column=col_num).fill = s["fill_navy"]
        for res in self.results:
            if res["status"] == "FAIL":
                ws3.append([res["test_id"], res["module"], res["name"], res["status"], res["exec_time_ms"], res["priority"], res["expected"], res["actual"], res.get("failure_reason", ""), res.get("screenshot", "")])

        # Sheet 4: Skipped Tests
        ws4 = wb.create_sheet(title="Skipped Tests")
        ws4.views.sheetView[0].showGridLines = True
        ws4.append(headers)
        for col_num, h_text in enumerate(headers, 1):
            ws4.cell(row=1, column=col_num).font = s["header"]
            ws4.cell(row=1, column=col_num).fill = s["fill_navy"]
        for res in self.results:
            if res["status"] not in ["PASS", "FAIL"]:
                ws4.append([res["test_id"], res["module"], res["name"], res["status"], res["exec_time_ms"], res["priority"], res["expected"], res["actual"]])

        # Sheet 5: Execution Metrics
        ws5 = wb.create_sheet(title="Execution Metrics")
        ws5.views.sheetView[0].showGridLines = True
        ws5.append(["Metric Name", "Value"])
        ws5.cell(row=1, column=1).font = s["header"]
        ws5.cell(row=1, column=1).fill = s["fill_navy"]
        ws5.cell(row=1, column=2).font = s["header"]
        ws5.cell(row=1, column=2).fill = s["fill_navy"]

        total = len(self.results)
        passed = sum(1 for r in self.results if r["status"] == "PASS")
        failed = sum(1 for r in self.results if r["status"] == "FAIL")
        skipped = total - (passed + failed)
        pass_pct = (passed / total * 100) if total > 0 else 0.0

        metrics_data = [
            ["Total Test Cases Designed", total],
            ["Total Test Cases Executed", total],
            ["Passed Test Cases", passed],
            ["Failed Test Cases", failed],
            ["Skipped / Blocked Test Cases", skipped],
            ["Overall Pass Percentage", f"{pass_pct:.2f}%"],
            ["Target Environment URL", Config.BASE_URL]
        ]
        for row in metrics_data:
            ws5.append(row)

        # Sheet 6: Defect Summary
        ws6 = wb.create_sheet(title="Defect Summary")
        ws6.views.sheetView[0].showGridLines = True
        ws6.append(["Defect ID", "Test Case ID", "Module", "Severity", "Failure Reason", "Status"])
        for col_num in range(1, 7):
            ws6.cell(row=1, column=col_num).font = s["header"]
            ws6.cell(row=1, column=col_num).fill = s["fill_navy"]

        defect_count = 1
        for res in self.results:
            if res["status"] == "FAIL":
                ws6.append([
                    f"DEF-SEL-{defect_count:03d}", res["test_id"], res["module"],
                    res["priority"], res.get("failure_reason", "Assertion failure"), "Open"
                ])
                defect_count += 1

        # Auto-adjust column widths
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 45)

        filepath = os.path.join(Config.EXCEL_DIR, "Automation_Test_Report.xlsx")
        wb.save(filepath)

    def _generate_passed_tests_report(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Passed Test Cases"
        ws.append(["Test ID", "Module", "Test Name", "Priority", "Execution Time (ms)"])
        for res in self.results:
            if res["status"] == "PASS":
                ws.append([res["test_id"], res["module"], res["name"], res["priority"], res["exec_time_ms"]])
        filepath = os.path.join(Config.EXCEL_DIR, "Passed_Test_Cases.xlsx")
        wb.save(filepath)

    def _generate_failed_tests_report(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Failed Test Cases"
        ws.append(["Test ID", "Module", "Test Name", "Priority", "Failure Reason", "Screenshot"])
        for res in self.results:
            if res["status"] == "FAIL":
                ws.append([res["test_id"], res["module"], res["name"], res["priority"], res.get("failure_reason", ""), res.get("screenshot", "")])
        filepath = os.path.join(Config.EXCEL_DIR, "Failed_Test_Cases.xlsx")
        wb.save(filepath)

    def _generate_summary_report(self):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Execution Summary"
        total = len(self.results)
        passed = sum(1 for r in self.results if r["status"] == "PASS")
        failed = sum(1 for r in self.results if r["status"] == "FAIL")
        skipped = total - (passed + failed)
        pass_rate = (passed / total * 100) if total > 0 else 0

        ws.append(["Metric", "Value"])
        ws.append(["Total Executed", total])
        ws.append(["Passed", passed])
        ws.append(["Failed", failed])
        ws.append(["Skipped", skipped])
        ws.append(["Pass Rate", f"{pass_rate:.2f}%"])

        filepath = os.path.join(Config.EXCEL_DIR, "Summary_Report.xlsx")
        wb.save(filepath)
